# kalibrasi_app/gui/app_window.py

import customtkinter as ctk
import os
import json
import shutil
import numpy as np
from datetime import datetime
from tkinter import filedialog
try:
    import win32com.client
    IS_WINDOWS = True
except ImportError:
    IS_WINDOWS = False

from gui.widgets.file_selector import load_seed_file
from gui.widgets.channel_selector import ChannelSelector
from gui.widgets.digitizer_selector import DigitizerSelector
from gui.widgets.loading_indicator import LoadingIndicator
from gui.windows.admin_data_popup import AdminDataPopup
from gui.windows.digitizer_popup import DigitizerPopup
from gui.windows.frequency_domain_popup import FrequencyDomainPopup
from gui.plotting.plot_frame import PlotFrame
from modules.freq_detector import detect_frequency_boundaries, detect_dominant_frequency
from modules.amplitude_extractor import find_best_amplitude_pairs
from obspy import Trace
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from tkinter import messagebox

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

class KalibrasiApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Aplikasi Kalibrasi Seismometer")
        self.geometry("1200x700")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.after(100, self._maximize)

        # === Sidebar ===
        self.sidebar_scrollable = ctk.CTkScrollableFrame(self, width=280)
        self.sidebar_scrollable.pack(side="left", fill="y", padx=10, pady=10)

        # Tombol awal
        self.browse_button = ctk.CTkButton(self.sidebar_scrollable, text="Browse File SEED", command=lambda: load_seed_file(self))
        self.browse_button.pack(pady=10, fill="x")
        self.digitizer_selector = DigitizerSelector(self.sidebar_scrollable, command=self.save_selected_digitizer_to_excel)
        self.digitizer_selector.pack(pady=10, fill="x")
        self.admin_button = ctk.CTkButton(self.sidebar_scrollable, text="Data Administrasi", command=self.show_admin_popup)
        self.admin_button.pack(pady=10, fill="x")
        
        # Widget yang disembunyikan
        self.channel_selector = ChannelSelector(self.sidebar_scrollable)
        self.channel_selector.pack_forget()
        self.set_boundary_button = ctk.CTkButton(self.sidebar_scrollable, text="1. Set Boundary & Identifikasi Frekuensi", command=self.set_boundary)
        self.set_boundary_button.pack_forget()
        self.delete_mode = ctk.BooleanVar(value=False)
        self.delete_mode_checkbox = ctk.CTkCheckBox(self.sidebar_scrollable, text="Aktifkan Mode Hapus Area (Drag)", variable=self.delete_mode)
        self.delete_mode_checkbox.pack_forget()
        self.extract_button = ctk.CTkButton(self.sidebar_scrollable, text="2. Ekstrak Amplitudo", fg_color="green", hover_color="#006400", command=self.extract_amplitude)
        self.extract_button.pack_forget()
        self.cert_button = ctk.CTkButton(self.sidebar_scrollable, text="3. Buat Sertifikat", command=self.export_certificate)
        self.cert_button.pack_forget()

        # === Plot Frame ===
        self.plot_frame = PlotFrame(self, on_boundaries_deleted_in_range=self.delete_boundaries_in_range)
        self.plot_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        # === State Variables ===
        self.stream = None
        self.boundary_trace = None
        self.boundaries = []
        self.identified_segments = []
        self.latest_admin_data = {}
        self.latest_digitizer_data = {}
        self.latest_amplitude_data = {}
        self.latest_freq_states = {}

    def _maximize(self): self.state("zoomed")
    def on_closing(self): self.destroy(); self.quit()

    def reset_ui_to_initial_state(self):
        print("[INFO] Mereset UI ke kondisi awal.")
        self.channel_selector.pack_forget()
        self.set_boundary_button.pack_forget()
        self.delete_mode_checkbox.pack_forget()
        self.delete_mode.set(False)
        self.extract_button.pack_forget()
        self.cert_button.pack_forget()
        self.plot_frame.clear_plot()
        self.stream = None
        self.boundary_trace = None
        self.boundaries = []
        self.identified_segments = []
        self.latest_admin_data = {}
        self.latest_digitizer_data = {}
        self.latest_amplitude_data = {}
        self.latest_freq_states = {}

    def update_plot_selected_channels(self):
        self.channel_selector.pack(pady=10, fill="x")
        self.set_boundary_button.pack(pady=10, fill="x")
        
        all_keys = sorted(set(f"{tr.stats.station}.{tr.stats.location}.{tr.stats.channel}" for tr in self.stream))
        key_map = {"East-West": None, "North-South": None, "Up-Down": None}
        available_for_mapping = list(all_keys)
        ch_ending_map = {"Up-Down": ('Z', 'UD'), "North-South": ('N', 'NS'), "East-West": ('E', 'EW')}
        for label, endings in ch_ending_map.items():
            found_key = next((key for key in available_for_mapping if key.split('.')[-1].upper().endswith(endings)), None)
            if found_key: key_map[label] = found_key; available_for_mapping.remove(found_key)
        remaining_keys = iter(available_for_mapping)
        for label in key_map:
            if key_map[label] is None: key_map[label] = next(remaining_keys, "")
        selected_keys = [key_map["East-West"], key_map["North-South"], key_map["Up-Down"]]; selected_keys = [k for k in selected_keys if k]
        self.channel_selector.set_channels(all_keys, selected_keys)
        self.plot_frame.plot_stream(self.stream, self.channel_selector.get_all_selected())

    def show_admin_popup(self):
        existing_data = self.load_admin_data_from_excel()
        AdminDataPopup(self, on_save_callback=self.save_admin_data_to_excel, existing_data=existing_data)

    def set_boundary(self):
        loading = LoadingIndicator(self, "Mendeteksi & Mengidentifikasi...")
        self.update_idletasks()
        try:
            channel_key = self.channel_selector.get_selected_channels()
            if not channel_key: return
            try:
                station, location, channel = channel_key.split(".")
                trace = self.stream.select(station=station, location=location, channel=channel).merge(method=1)[0]
            except Exception as e:
                print(f"Error selecting trace: {e}"); return
            
            self.boundary_trace = trace
            self.boundaries = detect_frequency_boundaries(trace)
            self.plot_frame.add_boundaries(self.boundaries)
            
            self._identify_and_annotate_segments()
            
            self.set_boundary_button.configure(text="Reload Boundary & Identifikasi Ulang")
            self.delete_mode_checkbox.pack(pady=(10,5), anchor="w")
            self.extract_button.pack(pady=5, fill="x")
        finally:
            loading.stop()

    def _identify_and_annotate_segments(self):
        """
        Mengidentifikasi segmen frekuensi berdasarkan self.boundaries saat ini
        dan langsung menggambar anotasinya di plot.
        """
        if not self.boundary_trace: return

        # Tambahkan awal dan akhir sinyal ke daftar boundary untuk analisis lengkap
        start_time = 0
        end_time = self.boundary_trace.stats.npts / self.boundary_trace.stats.sampling_rate
        full_boundaries = sorted(list(set([start_time] + self.boundaries + [end_time])))

        standard_freqs = np.array([0.01, 0.02, 0.05, 0.1, 0.2, 0.5, 1, 2, 5, 10, 15, 20])
        minimum_duration_seconds = 50
        self.identified_segments.clear()

        print("\n--- Mengidentifikasi Ulang Segmen Frekuensi ---")
        for i in range(len(full_boundaries) - 1):
            t_start = full_boundaries[i]
            t_end = full_boundaries[i+1]

            if (t_end - t_start) < minimum_duration_seconds:
                continue

            segment = self.boundary_trace.slice(self.boundary_trace.stats.starttime + t_start, self.boundary_trace.stats.starttime + t_end)
            if len(segment.data) == 0: continue

            detected_freq = detect_dominant_frequency(segment.data, segment.stats.sampling_rate)
            classified_freq = standard_freqs[np.argmin(np.abs(standard_freqs - detected_freq))]
            
            self.identified_segments.append((t_start, t_end, classified_freq))
            print(f"Segmen ({t_start:.1f}s-{t_end:.1f}s): Teridentifikasi -> {classified_freq} Hz")

        self.plot_frame.add_frequency_annotations(self.identified_segments)

    def delete_boundaries_in_range(self, times_to_delete):
        times_set = set(times_to_delete)
        self.boundaries = [b for b in self.boundaries if b not in times_set]
        print(f"{len(times_set)} data boundary dihapus dari list.")
        
        # Panggil ulang identifikasi dan anotasi setelah menghapus
        self._identify_and_annotate_segments()

    def extract_amplitude(self):
        loading = LoadingIndicator(self, "Mengekstrak Amplitudo...")
        self.update_idletasks()
        try:
            if not self.identified_segments:
                messagebox.showwarning("Perhatian", "Tidak ada segmen teridentifikasi untuk diekstrak.")
                return
            all_results = {}
            points_to_plot = {}
            standard_freqs = set([0.01, 0.02, 0.05, 0.1, 0.2, 0.5, 1, 2, 5, 10, 15, 20])
            for t_start, t_end, freq in self.identified_segments:
                values_by_ch = {}
                for ch_key in self.channel_selector.get_all_selected():
                    if not ch_key: continue
                    station, loc, ch = ch_key.split(".")
                    trace_ch = self.stream.select(station=station, location=loc, channel=ch).merge(method=1)[0]
                    segment_ch = trace_ch.slice(trace_ch.stats.starttime + t_start, trace_ch.stats.starttime + t_end)
                    best_pairs_info = find_best_amplitude_pairs(segment_ch.data)
                    simple_pairs = [(p['peak']['value'], p['trough']['value']) for p in best_pairs_info]
                    simple_name = "UD" if ch.upper().endswith(('Z', 'UD')) else "NS" if ch.upper().endswith(('N', 'NS')) else "EW" if ch.upper().endswith(('E', 'EW')) else None
                    if simple_name: values_by_ch[simple_name] = simple_pairs
                    if ch_key not in points_to_plot: points_to_plot[ch_key] = []
                    fs = segment_ch.stats.sampling_rate
                    for pair in best_pairs_info:
                        points_to_plot[ch_key].append((t_start + (pair['peak']['index'] / fs), pair['peak']['value']))
                        points_to_plot[ch_key].append((t_start + (pair['trough']['index'] / fs), pair['trough']['value']))
                if freq in all_results:
                    for ch_name, pairs in values_by_ch.items(): all_results[freq].setdefault(ch_name, []).extend(pairs)
                else:
                    all_results[freq] = values_by_ch
            
            self.latest_amplitude_data = all_results
            self.latest_freq_states = {freq: (freq in all_results) for freq in standard_freqs}
            self.plot_frame.show_amplitude_table(all_results)
            self.save_amplitude_data_to_excel()
            self.plot_frame.plot_selected_points(points_to_plot)
            self.cert_button.pack(pady=20, fill="x")
            messagebox.showinfo("Sukses", "Ekstraksi amplitudo selesai.")
        finally:
            loading.stop()

    def export_certificate(self):
        source_excel_path = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        if not os.path.exists(source_excel_path):
            messagebox.showerror("Error", f"File data '{os.path.basename(source_excel_path)}' tidak ditemukan."); return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suggested_name = f"Sertifikat_Final_{timestamp}.pdf"
        
        pdf_path = filedialog.asksaveasfilename(
            title="Simpan Sertifikat Sebagai PDF",
            initialfile=suggested_name,
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if not pdf_path:
            print("Penyimpanan sertifikat dibatalkan.")
            return

        loading = LoadingIndicator(self, "Membuat Sertifikat...")
        self.update_idletasks()
        try:
            temp_excel_path = pdf_path.replace(".pdf", ".xlsx")
            shutil.copy(source_excel_path, temp_excel_path)
            
            wb = openpyxl.load_workbook(temp_excel_path)
            if 'SERTIF_SEISMO' not in wb.sheetnames:
                messagebox.showerror("Error", "Sheet 'SERTIF_SEISMO' tidak ditemukan."); wb.close(); return
            
            ws_cert = wb['SERTIF_SEISMO']
            plot_image_path = "data/plots/clean_plot.png"
            if os.path.exists(plot_image_path):
                img = Image(plot_image_path); img.width *= 0.90; img.height *= 0.50
                ws_cert.add_image(img, 'B136')
            
            wb.save(temp_excel_path)
            
            self._export_sheet_to_pdf(temp_excel_path, 'SERTIF_SEISMO', pdf_path)
            
            os.remove(temp_excel_path)
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat file sertifikat:\n{e}")
        finally:
            loading.stop()

    # ... Sisa file (fungsi helper) tidak berubah ...
    def load_admin_data_from_excel(self):
        filepath = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        if not os.path.exists(filepath): return {}
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "data_administrasi" not in wb.sheetnames: return {}
            ws = wb["data_administrasi"]; data = {}
            for row in ws.iter_rows(min_row=1, max_col=2):
                key_cell, value_cell = row
                if key_cell.value and value_cell.value is not None: data[str(key_cell.value)] = str(value_cell.value)
            return data
        except Exception as e:
            print(f"Gagal memuat data admin: {e}"); return {}
    def _get_or_create_workbook_and_sheet(self, filepath, sheet_name):
        try: wb = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
        if sheet_name in wb.sheetnames: wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        return wb, ws
    def save_admin_data_to_excel(self, data):
        self.latest_admin_data = data; output_file = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        wb, ws = self._get_or_create_workbook_and_sheet(output_file, 'data_administrasi')
        ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 50
        row = 1
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True); ws.cell(row=row, column=2, value=value); row += 1
        wb.save(output_file); print(f"Data admin disimpan ke {output_file}")
    def save_selected_digitizer_to_excel(self, selected_name):
        if selected_name == "Add New...": return
        try:
            with open("data/digitizer_config.json", "r") as f: digitizers = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError): return
        selected_data = next((d for d in digitizers if d.get("Name") == selected_name), None)
        if not selected_data: return
        self.latest_digitizer_data = selected_data; output_file = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        wb, ws = self._get_or_create_workbook_and_sheet(output_file, 'data_digitizer')
        ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 40
        row = 1
        for key, value in selected_data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True); ws.cell(row=row, column=2, value=value); row += 1
        wb.save(output_file); print(f"Data digitizer disimpan ke {output_file}")
    def _write_amplitude_data_to_sheet(self, ws, processed_data, all_freq_states):
        headers = ["FREKUENSI (Hz)", "NS Max", "NS Min", "EW Max", "EW Min", "UD Max", "UD Min"]
        for col, h in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col, value=h); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
            if "NS" in h: cell.fill = PatternFill(fill_type="solid", start_color="BDD7EE")
            elif "EW" in h: cell.fill = PatternFill(fill_type="solid", start_color="FFE699")
            elif "UD" in h: cell.fill = PatternFill(fill_type="solid", start_color="C6E0B4")
        row_idx = 5
        for freq in sorted(all_freq_states.keys()):
            is_enabled = all_freq_states.get(freq, False)
            ws.cell(row=row_idx, column=2, value=float(freq)).font = Font(bold=True)
            if is_enabled and freq in processed_data:
                ch_data = processed_data.get(freq, {}); ns_pairs = ch_data.get("NS", []); ew_pairs = ch_data.get("EW", []); ud_pairs = ch_data.get("UD", []); max_rows_for_freq = max(len(ns_pairs), len(ew_pairs), len(ud_pairs), 1)
                for i in range(max_rows_for_freq):
                    current_row = row_idx + i
                    if i < len(ns_pairs): ws.cell(row=current_row, column=3, value=ns_pairs[i][0]); ws.cell(row=current_row, column=4, value=ns_pairs[i][1])
                    if i < len(ew_pairs): ws.cell(row=current_row, column=5, value=ew_pairs[i][0]); ws.cell(row=current_row, column=6, value=ew_pairs[i][1])
                    if i < len(ud_pairs): ws.cell(row=current_row, column=7, value=ud_pairs[i][0]); ws.cell(row=current_row, column=8, value=ud_pairs[i][1])
                row_idx += max_rows_for_freq
            else: row_idx += 1
    def save_amplitude_data_to_excel(self):
        output_file = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        wb, ws = self._get_or_create_workbook_and_sheet(output_file, 'data_auto')
        self._write_amplitude_data_to_sheet(ws, self.latest_amplitude_data, self.latest_freq_states)
        wb.save(output_file); print(f"Data amplitudo disimpan ke {output_file}")
    def _export_sheet_to_pdf(self, excel_path, sheet_name, pdf_path):
        if not IS_WINDOWS:
            messagebox.showerror("Error", "Ekspor PDF otomatis hanya di Windows dengan Excel."); return False
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application"); excel.Visible = False; excel_path_abs = os.path.abspath(excel_path); pdf_path_abs = os.path.abspath(pdf_path)
            workbook = excel.Workbooks.Open(excel_path_abs); workbook.Worksheets(sheet_name).Activate(); workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_path_abs)
            messagebox.showinfo("Sukses", f"Berhasil ekspor sertifikat ke:\n{pdf_path_abs}"); return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal ekspor ke PDF:\n{e}\n\nPastikan MS Excel terinstal & sheet '{sheet_name}' ada."); return False
        finally:
            if excel and hasattr(excel, 'ActiveWorkbook') and excel.ActiveWorkbook: excel.ActiveWorkbook.Close(SaveChanges=False)
            if excel: excel.Quit()

if __name__ == '__main__':
    app = KalibrasiApp()
    app.mainloop()


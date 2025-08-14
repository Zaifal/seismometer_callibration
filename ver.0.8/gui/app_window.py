# kalibrasi_app/gui/app_window.py

import customtkinter as ctk
import os
import json
import shutil
from datetime import datetime
try:
    # Library ini hanya ada di Windows dan diperlukan untuk ekspor PDF
    import win32com.client
    IS_WINDOWS = True
except ImportError:
    IS_WINDOWS = False

from gui.widgets.file_selector import load_seed_file
from gui.widgets.channel_selector import ChannelSelector
from gui.widgets.digitizer_selector import DigitizerSelector
from gui.windows.admin_data_popup import AdminDataPopup
from gui.windows.digitizer_popup import DigitizerPopup
from gui.windows.freq_selector_frame import FreqSelectorFrame
from gui.plotting.plot_frame import PlotFrame
from modules.freq_detector import detect_amplitude_boundaries
from modules.amplitude_extractor import find_best_amplitude_pairs
from obspy import Trace
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from tkinter import messagebox

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

class KalibrasiApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Aplikasi Kalibrasi Seismometer")
        self.geometry("1200x700")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.after(100, self._maximize)

        # === Bagian Sidebar (GUI) ===
        self.sidebar_scrollable = ctk.CTkScrollableFrame(self, width=280)
        self.sidebar_scrollable.pack(side="left", fill="y", padx=10, pady=10)

        self.browse_button = ctk.CTkButton(self.sidebar_scrollable, text="Browse File SEED", command=lambda: load_seed_file(self))
        self.browse_button.pack(pady=10, fill="x")
        
        self.digitizer_selector = DigitizerSelector(
            self.sidebar_scrollable,
            command=self.save_selected_digitizer_to_excel
        )
        self.digitizer_selector.pack(pady=10, fill="x")

        self.admin_button = ctk.CTkButton(self.sidebar_scrollable, text="Data Administrasi", command=self.show_admin_popup)
        self.admin_button.pack(pady=10, fill="x")

        self.channel_selector = ChannelSelector(self.sidebar_scrollable)
        self.channel_selector.pack(pady=10, fill="x")
        self.channel_selector.hide()

        self.threshold_label = ctk.CTkLabel(self.sidebar_scrollable, text="Threshold Ratio Boundary")
        self.threshold_entry = ctk.CTkEntry(self.sidebar_scrollable, placeholder_text="e.g. 0.5")
        self.threshold_entry.insert(0, "0.5")
        self.threshold_label.pack_forget()
        self.threshold_entry.pack_forget()

        self.set_boundary_button = ctk.CTkButton(self.sidebar_scrollable, text="1. Set Boundary", command=self.set_boundary)
        self.set_boundary_button.pack(pady=10, fill="x")
        self.set_boundary_button.pack_forget()

        self.freq_selector = FreqSelectorFrame(self.sidebar_scrollable, on_add_click=self.associate_clicks_to_frequency)
        self.freq_selector.pack(pady=10, fill="x")
        self.freq_selector.pack_forget()

        self.extract_button = ctk.CTkButton(self.sidebar_scrollable, text="2. Ekstrak Amplitudo", command=self.extract_amplitude_and_update)
        self.extract_button.pack(pady=10, fill="x")
        self.extract_button.pack_forget()
        
        self.cert_button = ctk.CTkButton(self.sidebar_scrollable, text="3. Buat Sertifikat", command=self.export_certificate)
        self.cert_button.pack(pady=20, fill="x")
        self.cert_button.pack_forget()


        # === Bagian Plot Frame ===
        self.plot_frame = PlotFrame(self, on_new_letter_callback=self.handle_new_letter, on_letter_removed_callback=self.handle_remove_letter, on_clear_all_callback=self.reset_all_associations)
        self.plot_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        # === Variabel State Aplikasi ===
        self.stream = None
        self.boundary_trace = None
        self.boundaries = []
        self.all_letter_clicks = {}
        self.freq_letter_map = {}
        self.associated_letters = set()
        self.latest_admin_data = {}
        self.latest_digitizer_data = {}
        self.latest_amplitude_data = {}
        self.latest_freq_states = {}

    def _maximize(self):
        self.state("zoomed")

    def on_closing(self):
        self.destroy()
        self.quit()

    def reset_ui_to_initial_state(self):
        print("[INFO] Mereset UI ke kondisi awal.")
        
        self.channel_selector.hide()
        self.threshold_label.pack_forget()
        self.threshold_entry.pack_forget()
        self.set_boundary_button.pack_forget()
        self.freq_selector.pack_forget()
        self.extract_button.pack_forget()
        self.cert_button.pack_forget()

        self.plot_frame.clear_plot()
        self.reset_all_associations()

        self.stream = None
        self.boundary_trace = None
        self.boundaries = []
        self.latest_admin_data = {}
        self.latest_digitizer_data = {}
        self.latest_amplitude_data = {}
        self.latest_freq_states = {}

    def update_plot_selected_channels(self):
        all_keys = sorted(set(f"{tr.stats.station}.{tr.stats.location}.{tr.stats.channel}" for tr in self.stream))
        key_map = {"East-West": None, "North-South": None, "Up-Down": None}
        available_for_mapping = list(all_keys)
        ch_ending_map = {"Up-Down": ('Z', 'UD'), "North-South": ('N', 'NS'), "East-West": ('E', 'EW')}
        
        for label, endings in ch_ending_map.items():
            found_key = next((key for key in available_for_mapping if key.split('.')[-1].upper().endswith(endings)), None)
            if found_key:
                key_map[label] = found_key
                available_for_mapping.remove(found_key)
        
        remaining_keys = iter(available_for_mapping)
        for label in key_map:
            if key_map[label] is None:
                key_map[label] = next(remaining_keys, "")
                
        selected_keys = [key_map["East-West"], key_map["North-South"], key_map["Up-Down"]]
        selected_keys = [k for k in selected_keys if k]
        
        self.channel_selector.set_channels(all_keys, selected_keys)
        self.channel_selector.show()
        self.threshold_label.pack(pady=(10, 0))
        self.threshold_entry.pack(pady=(0, 10))
        self.set_boundary_button.pack(pady=10, fill="x")
        self.plot_frame.plot_stream(self.stream, self.channel_selector.get_all_selected())

    def load_admin_data_from_excel(self):
        """Membaca data administrasi yang ada dari file Excel."""
        filepath = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        if not os.path.exists(filepath):
            return {}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "data_administrasi" not in wb.sheetnames:
                return {}
            
            ws = wb["data_administrasi"]
            data = {}
            for row in ws.iter_rows(min_row=1, max_col=2):
                key_cell, value_cell = row
                if key_cell.value and value_cell.value is not None:
                    data[str(key_cell.value)] = str(value_cell.value)
            return data
        except Exception as e:
            print(f"Gagal memuat data administrasi dari Excel: {e}")
            return {}

    def show_admin_popup(self):
        # 1. Muat data yang sudah ada dari Excel
        existing_data = self.load_admin_data_from_excel()
        
        # 2. Kirim data tersebut saat membuat popup
        AdminDataPopup(self, 
                       on_save_callback=self.save_admin_data_to_excel,
                       existing_data=existing_data)

    def set_boundary(self):
        try:
            threshold = float(self.threshold_entry.get())
        except ValueError:
            threshold = 0.5
        
        channel_key = self.channel_selector.get_selected_channels()
        if not channel_key: return
        
        try:
            station, location, channel = channel_key.split(".")
            trace = self.stream.select(station=station, location=location, channel=channel).merge(method=1)[0]
        except Exception as e:
            print(f"Error selecting trace for boundary: {e}")
            return
            
        self.boundary_trace = trace
        self.boundaries = detect_amplitude_boundaries(trace, threshold_ratio=threshold)
        self.plot_frame.plot_trace_with_boundaries(trace, self.boundaries)
        self.set_boundary_button.configure(text="Reload Boundary")
        self.freq_selector.pack(pady=10, fill="x")
        self.extract_button.pack(pady=10, fill="x")
        self.cert_button.pack(pady=20, fill="x")
        self.reset_all_associations()

    def handle_new_letter(self, letter, time):
        self.all_letter_clicks[letter] = time

    def associate_clicks_to_frequency(self, freq):
        current_letters_on_plot = set(self.all_letter_clicks.keys())
        newly_associated_letters = sorted(list(current_letters_on_plot - self.associated_letters))
        
        if not newly_associated_letters:
            messagebox.showwarning("Perhatian", f"Tidak ada huruf baru pada plot untuk ditambahkan ke frekuensi {freq}")
            return
            
        self.freq_letter_map[freq] = newly_associated_letters
        self.associated_letters.update(newly_associated_letters)
        self.freq_selector.display_letters_for_freq(freq, newly_associated_letters)
        print(f"[INFO] Frekuensi {freq} Hz terhubung dengan huruf: {newly_associated_letters}")

    def handle_remove_letter(self, letter):
        if letter in self.all_letter_clicks: del self.all_letter_clicks[letter]
        if letter in self.associated_letters: self.associated_letters.remove(letter)
        
        for freq, letters in self.freq_letter_map.items():
            if letter in letters:
                letters.remove(letter)
                self.freq_selector.display_letters_for_freq(freq, letters)
                break
    
    def reset_all_associations(self):
        self.all_letter_clicks.clear()
        self.freq_letter_map.clear()
        self.associated_letters.clear()
        if hasattr(self, 'freq_selector'):
            for freq in self.freq_selector.freq_widgets.keys():
                self.freq_selector.display_letters_for_freq(freq, [])

    def _get_or_create_workbook_and_sheet(self, filepath, sheet_name):
        try:
            wb = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        return wb, ws

    def save_admin_data_to_excel(self, data):
        self.latest_admin_data = data
        output_file = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        wb, ws = self._get_or_create_workbook_and_sheet(output_file, 'data_administrasi')
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        row = 1
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        wb.save(output_file)
        print(f"Data administrasi disimpan ke sheet 'data_administrasi' di {output_file}")

    def save_selected_digitizer_to_excel(self, selected_name):
        if selected_name == "Add New...": return
        
        try:
            with open("data/digitizer_config.json", "r") as f:
                digitizers = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return
            
        selected_data = next((d for d in digitizers if d.get("Name") == selected_name), None)
        if not selected_data: return
        
        self.latest_digitizer_data = selected_data
        output_file = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        wb, ws = self._get_or_create_workbook_and_sheet(output_file, 'data_digitizer')
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        row = 1
        for key, value in selected_data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        wb.save(output_file)
        print(f"Data digitizer '{selected_name}' disimpan ke sheet 'data_digitizer' di {output_file}")

    def extract_amplitude_and_update(self):
        self.latest_freq_states = self.freq_selector.get_all_freq_states()
        output_by_freq = {}
        points_to_plot = {}

        if not self.boundaries:
            messagebox.showerror("Error", "Boundaries belum di-set.")
            return

        for freq, is_enabled in self.latest_freq_states.items():
            if not is_enabled: continue
            
            associated_letters = self.freq_letter_map.get(freq, [])
            if not associated_letters: continue
            
            letter_times = [self.all_letter_clicks[ltr] for ltr in associated_letters if ltr in self.all_letter_clicks]
            if not letter_times: continue
            
            first_time = min(letter_times)
            last_time = max(letter_times)

            try:
                t_start = max(t for t in self.boundaries if t <= first_time)
                t_end = min(t for t in self.boundaries if t >= last_time)
            except ValueError:
                print(f"[ERROR] Tidak dapat menemukan boundary untuk freq {freq}.")
                continue
            
            values_by_ch = {}
            for ch_key in self.channel_selector.get_all_selected():
                if not ch_key: continue
                
                station, loc, ch = ch_key.split(".")
                trace = self.stream.select(station=station, location=loc, channel=ch).merge(method=1)[0]
                segment = trace.slice(starttime=trace.stats.starttime + t_start, endtime=trace.stats.starttime + t_end)
                
                best_pairs_with_indices = find_best_amplitude_pairs(segment.data, max_pairs=5)
                
                simple_pairs = [(p['peak']['value'], p['trough']['value']) for p in best_pairs_with_indices]
                simple_name = "UD" if ch.upper().endswith(('Z', 'UD')) else "NS" if ch.upper().endswith(('N', 'NS')) else "EW" if ch.upper().endswith(('E', 'EW')) else None
                if simple_name:
                    values_by_ch[simple_name] = simple_pairs

                if ch_key not in points_to_plot:
                    points_to_plot[ch_key] = []
                
                fs = segment.stats.sampling_rate
                for pair in best_pairs_with_indices:
                    peak_time = t_start + (pair['peak']['index'] / fs)
                    trough_time = t_start + (pair['trough']['index'] / fs)
                    points_to_plot[ch_key].append((peak_time, pair['peak']['value']))
                    points_to_plot[ch_key].append((trough_time, pair['trough']['value']))
            
            output_by_freq[freq] = values_by_ch
            
        self.latest_amplitude_data = output_by_freq
        self.plot_frame.show_amplitude_table(self.latest_amplitude_data)
        self.save_amplitude_data_to_excel()
        self.plot_frame.plot_selected_points(points_to_plot)
        messagebox.showinfo("Sukses", "Ekstraksi amplitudo selesai dan titik sampel ditampilkan.")
        
    def _write_amplitude_data_to_sheet(self, ws, processed_data, all_freq_states):
        headers = ["FREKUENSI (Hz)", "NS Max", "NS Min", "EW Max", "EW Min", "UD Max", "UD Min"]
        for col, h in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if "NS" in h: cell.fill = PatternFill(fill_type="solid", start_color="BDD7EE")
            elif "EW" in h: cell.fill = PatternFill(fill_type="solid", start_color="FFE699")
            elif "UD" in h: cell.fill = PatternFill(fill_type="solid", start_color="C6E0B4")
            
        row_idx = 5
        for freq in sorted(all_freq_states.keys()):
            is_enabled = all_freq_states[freq]
            ws.cell(row=row_idx, column=2, value=float(freq)).font = Font(bold=True)
            if is_enabled and freq in processed_data:
                ch_data = processed_data.get(freq, {})
                ns_pairs = ch_data.get("NS", [])
                ew_pairs = ch_data.get("EW", [])
                ud_pairs = ch_data.get("UD", [])
                max_rows_for_freq = max(len(ns_pairs), len(ew_pairs), len(ud_pairs), 1)
                for i in range(max_rows_for_freq):
                    current_row = row_idx + i
                    if i < len(ns_pairs):
                        ws.cell(row=current_row, column=3, value=ns_pairs[i][0])
                        ws.cell(row=current_row, column=4, value=ns_pairs[i][1])
                    if i < len(ew_pairs):
                        ws.cell(row=current_row, column=5, value=ew_pairs[i][0])
                        ws.cell(row=current_row, column=6, value=ew_pairs[i][1])
                    if i < len(ud_pairs):
                        ws.cell(row=current_row, column=7, value=ud_pairs[i][0])
                        ws.cell(row=current_row, column=8, value=ud_pairs[i][1])
                row_idx += max_rows_for_freq
            else:
                row_idx += 1
    
    def save_amplitude_data_to_excel(self):
        output_file = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        wb, ws = self._get_or_create_workbook_and_sheet(output_file, 'data_auto')
        self._write_amplitude_data_to_sheet(ws, self.latest_amplitude_data, self.latest_freq_states)
        wb.save(output_file)
        print(f"Data amplitudo disimpan ke sheet 'data_auto' di {output_file}")

    def _export_sheet_to_pdf(self, excel_path, sheet_name, pdf_path):
        if not IS_WINDOWS:
            messagebox.showerror("Error", "Fitur ekspor PDF otomatis hanya berfungsi di Windows dengan Excel terinstal.")
            return False
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel_path_abs = os.path.abspath(excel_path)
            pdf_path_abs = os.path.abspath(pdf_path)
            workbook = excel.Workbooks.Open(excel_path_abs)
            
            workbook.Worksheets(sheet_name).Activate()
            
            workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_path_abs)
            
            messagebox.showinfo("Sukses", f"Berhasil mengekspor sertifikat ke:\n{pdf_path_abs}")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal mengekspor ke PDF:\n{e}\n\nPastikan MS Excel terinstal dan sheet '{sheet_name}' ada di file Excel.")
            return False
        finally:
            if excel and hasattr(excel, 'ActiveWorkbook') and excel.ActiveWorkbook:
                excel.ActiveWorkbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()

    def export_certificate(self):
        source_excel_path = os.path.join("data", "amplitudo_ekstraksi.xlsx")
        
        if not os.path.exists(source_excel_path):
            messagebox.showerror("Error", f"File '{os.path.basename(source_excel_path)}' tidak ditemukan.\nHarap lakukan 'Ekstrak Amplitudo' setidaknya sekali.")
            return
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_excel_name = f"Sertifikat_Output_{timestamp}.xlsx"
        output_pdf_name = f"Sertifikat_Final_{timestamp}.pdf"
        output_excel_path = os.path.join("data", output_excel_name)
        output_pdf_path = os.path.join("data", output_pdf_name)

        try:
            shutil.copy(source_excel_path, output_excel_path)
            print(f"File kerja disalin ke: {output_excel_path}")

            wb = openpyxl.load_workbook(output_excel_path)

            if 'SERTIF_SEISMO' not in wb.sheetnames:
                messagebox.showerror("Error", "Sheet 'SERTIF_SEISMO' tidak ditemukan di dalam file Excel.")
                wb.close()
                return

            ws_cert = wb['SERTIF_SEISMO']
            
            plot_image_path = "data/plots/clean_plot.png"
            if os.path.exists(plot_image_path):
                img = Image(plot_image_path)
                
                img.width = img.width * 0.90
                img.height = img.height * 0.50
                
                ws_cert.add_image(img, 'B136')
                print(f"Gambar plot dimasukkan ke sel B136.")
            else:
                print(f"[WARNING] File gambar plot tidak ditemukan di {plot_image_path}")

            wb.save(output_excel_path)

            self._export_sheet_to_pdf(output_excel_path, 'SERTIF_SEISMO', output_pdf_path)

        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat file sertifikat:\n{e}")

if __name__ == '__main__':
    app = KalibrasiApp()
    app.mainloop()